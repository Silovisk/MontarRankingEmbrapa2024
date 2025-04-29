#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Script to parse EMBRAPA exam results and build a ranking
Date: April 29, 2025
"""

import re
import os
import time
import pandas as pd
from operator import itemgetter
from tabulate import tabulate
import sys
import colorama
from colorama import Fore, Back, Style
import csv

# Initialize colorama for Windows terminal colors
colorama.init()

def print_progress_bar(iteration, total, prefix='', suffix='', length=50, fill='█', print_end='\n'):
    """
    Call in a loop to create terminal progress bar with color
    """
    percent = ("{0:.1f}").format(100 * (iteration / float(total)))
    filled_length = int(length * iteration // total)
    bar = fill * filled_length + '-' * (length - filled_length)
    
    # Add colors for better visibility
    color_prefix = Fore.CYAN + prefix + Style.RESET_ALL
    color_percent = Fore.GREEN + percent + "%" + Style.RESET_ALL
    color_bar = Fore.BLUE + "|" + Fore.YELLOW + bar + Fore.BLUE + "|" + Style.RESET_ALL
    color_suffix = Fore.CYAN + suffix + Style.RESET_ALL
    
    print(f'\r{color_prefix} {color_bar} {color_percent} {color_suffix}', end=print_end)
    sys.stdout.flush()

def show_spinner(seconds, message):
    """
    Display a spinner animation for the specified number of seconds
    """
    spinner = ['-', '\\', '|', '/']
    start_time = time.time()
    i = 0
    
    # Reduced spinner time for speed
    max_time = min(seconds, 0.2)  # Cap at 0.2 seconds for faster execution
    
    while time.time() - start_time < max_time:
        sys.stdout.write(f'\r{Fore.CYAN}{message} {Fore.YELLOW}{spinner[i % len(spinner)]}{Style.RESET_ALL}')
        sys.stdout.flush()
        time.sleep(0.02)  # Reduced sleep time
        i += 1
    
    sys.stdout.write('\r' + ' ' * (len(message) + 10) + '\r')
    sys.stdout.flush()

def parse_candidates_data(filename):
    """
    Parse the file content to extract candidate data in a format suitable for ranking
    """
    try:
        # Read the file content
        print(f"{Fore.CYAN}Reading file content...{Style.RESET_ALL}")
        show_spinner(0.2, "Opening and reading file...")
        
        with open(filename, 'r', encoding='utf-8') as file:
            content = file.read()
        
        # Extract option title and number
        show_spinner(0.1, "Extracting option information...")
        
        title_pattern = r'OPÇÃO (\d+): (.+)'
        title_match = re.search(title_pattern, content)
        option_number = title_match.group(1) if title_match else "Unknown"
        option_title = title_match.group(2) if title_match else "Unknown"
        
        print(f"{Fore.GREEN}Extracted option: {option_number} - {option_title}{Style.RESET_ALL}")
        
        # Clean up content for better parsing
        show_spinner(0.1, "Cleaning up data...")
        
        # Primeiro, vamos lidar com o problema do espaço no número decimal
        content = re.sub(r'(\d+)\.\s+(\d+)', r'\1.\2', content)
        
        # Remover quebras de linha e espaços extras
        cleaned_content = content.replace('\n', ' ').replace('  ', ' ')
        
        # Split by "/" delimiter which separates candidates
        candidate_chunks = cleaned_content.split('/')
        
        # Process each chunk to extract candidate data
        entries = []
        candidates_data_started = False
        
        for chunk in candidate_chunks:
            chunk = chunk.strip()
            
            # Skip empty chunks
            if not chunk:
                continue
                
            # If this is the first chunk, we need to extract the candidate part after the option title
            if not candidates_data_started and 'OPÇÃO' in chunk:
                parts = chunk.split('SUBÁREA:')
                if len(parts) > 1:
                    chunk = parts[1].strip()
                    candidates_data_started = True
                else:
                    continue
            else:
                candidates_data_started = True
            
            # Extract raw data using regex with more flexibility
            pattern = r'(\d+),\s*([\w\s.\-\']+),\s*(\d+\.\d+),\s*(\d+),\s*(\d+\.\d+),\s*(\d+),\s*(\d+\.\d+)'
            match = re.search(pattern, chunk)
            
            if match:
                entries.append(match.groups())
            else:
                print(f"{Fore.YELLOW}Trying alternative parsing for: {chunk}{Style.RESET_ALL}")
                
                # Manual parsing as fallback
                parts = [p.strip() for p in chunk.split(',')]
                if len(parts) >= 7:
                    # Fix any additional formatting issues in the parts
                    for i in range(2, 7):  # Process the numeric fields
                        if i < len(parts):
                            parts[i] = parts[i].replace(' ', '')
                    
                    try:
                        # Validate numeric fields
                        registration = parts[0]
                        name = parts[1]
                        p1_score = float(parts[2])
                        p1_correct = int(parts[3])
                        p2_score = float(parts[4])
                        p2_correct = int(parts[5])
                        final_score = float(parts[6])
                        
                        entries.append((registration, name, str(p1_score), str(p1_correct), 
                                        str(p2_score), str(p2_correct), str(final_score)))
                        print(f"{Fore.GREEN}Successfully parsed candidate using manual method{Style.RESET_ALL}")
                    except (ValueError, IndexError) as e:
                        print(f"{Fore.RED}Failed to parse: {e} - {parts}{Style.RESET_ALL}")
        
        print(f"{Fore.GREEN}✓ Found {len(entries)} candidates{Style.RESET_ALL}")
        
        # Process each candidate entry
        candidates = []
        print(f"\n{Fore.YELLOW}⏳ Processing candidates:{Style.RESET_ALL}")
        total_candidates = len(entries)
        
        for i, entry in enumerate(entries):
            try:
                # Update progress bar
                print_progress_bar(i + 1, total_candidates, 
                                 prefix=f'{Fore.CYAN}Progress:{Style.RESET_ALL}', 
                                 suffix=f'Complete ({i+1}/{total_candidates})', 
                                 print_end='\r')
                
                registration = entry[0].strip()
                name = entry[1].strip()
                p1_score = float(entry[2])
                p1_correct = int(entry[3])
                p2_score = float(entry[4])
                p2_correct = int(entry[5])
                final_score = float(entry[6])
                
                candidate = {
                    'registration': registration,
                    'name': name,
                    'p1_score': p1_score,
                    'p1_correct': p1_correct,
                    'p2_score': p2_score,
                    'p2_correct': p2_correct,
                    'final_score': final_score
                }
                candidates.append(candidate)
                
            except (IndexError, ValueError) as e:
                print(f"\n{Fore.RED}Error processing candidate: {e}{Style.RESET_ALL}")
                print(f"{Fore.RED}Problematic data: {entry}{Style.RESET_ALL}")
        
        print(f"\n{Fore.GREEN}✓ Successfully processed {len(candidates)} candidates{Style.RESET_ALL}")
        return option_number, option_title, candidates
        
    except Exception as e:
        print(f"{Fore.RED}Error processing file: {e}{Style.RESET_ALL}")
        return "Unknown", "Unknown", []

def build_ranking(candidates):
    """
    Sort candidates by final score in descending order
    """
    print(f"{Fore.YELLOW}⏳ Sorting candidates by final score...{Style.RESET_ALL}")
    
    # Show a sorting progress bar - faster animation
    steps = 5
    for i in range(steps + 1):
        print_progress_bar(i, steps, 
                          prefix=f'{Fore.CYAN}Sorting:{Style.RESET_ALL}', 
                          suffix=f'Complete', 
                          print_end='\r')
        time.sleep(0.01)
    
    ranked_candidates = sorted(candidates, key=itemgetter('final_score'), reverse=True)
    print(f"\n{Fore.GREEN}✓ Ranking completed successfully{Style.RESET_ALL}")
    return ranked_candidates

def format_excel(df, writer, option_number, option_title, total_candidates):
    """
    Format the Excel file with headers, colors, and proper column widths
    """
    # Import the necessary module
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Write the dataframe to Excel
    df.to_excel(writer, sheet_name='Ranking', startrow=6, index=False)
    
    # Access the workbook and the worksheet
    workbook = writer.book
    worksheet = writer.sheets['Ranking']
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    title_font = Font(name='Arial', size=14, bold=True)
    subtitle_font = Font(name='Arial', size=12, bold=True)
    normal_font = Font(name='Arial', size=10)
    
    # Add headers to the Excel file
    worksheet['A1'] = 'EMPRESA BRASILEIRA DE PESQUISA AGROPECUÁRIA (EMBRAPA)'
    worksheet['A2'] = 'CONCURSO PÚBLICO - EDITAL Nº 17 – EMBRAPA, DE 28 DE ABRIL DE 2025'
    worksheet['A4'] = f'OPÇÃO {option_number}: {option_title}'
    worksheet['A5'] = f'Total Candidates: {total_candidates}'
    
    # Format headers
    worksheet['A1'].font = title_font
    worksheet['A2'].font = normal_font
    worksheet['A4'].font = subtitle_font
    worksheet['A5'].font = normal_font
    
    # Format the table headers (row 7)
    for col_num, column_title in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_num)
        cell = worksheet[f'{col_letter}7']
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Format table
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Apply borders and center alignment to data
    for row_num in range(7, len(df) + 8):  # +7 for header rows, +1 for 0-index
        for col_num in range(1, len(df.columns) + 1):
            col_letter = get_column_letter(col_num)
            cell = worksheet[f'{col_letter}{row_num}']
            cell.border = thin_border
            
            # Center numeric columns
            if col_num != 3:  # Not the Name column
                cell.alignment = Alignment(horizontal='center')
    
    # Set column widths
    column_widths = {
        'Rank': 8,
        'Registration': 15, 
        'Name': 35,
        'P1 Score': 12,
        'P1 Correct': 12,
        'P2 Score': 12,
        'P2 Correct': 12,
        'Final Score': 12
    }
    
    for i, column in enumerate(df.columns, 1):
        col_letter = get_column_letter(i)
        worksheet.column_dimensions[col_letter].width = column_widths.get(column, 15)
    
    # Merge cells for headers
    worksheet.merge_cells('A1:H1')
    worksheet.merge_cells('A2:H2')
    worksheet.merge_cells('A4:H4')
    worksheet.merge_cells('A5:H5')
    
    return workbook

def display_ranking(option_number, option_title, candidates, output_format='csv', output_file=None):
    """
    Display the ranking in a formatted table and save to file in requested format
    """
    if not candidates:
        print(f"{Fore.RED}No candidates found.{Style.RESET_ALL}")
        return
    
    # Create a DataFrame with our data for better presentation - Faster progress
    print(f"{Fore.YELLOW}⏳ Preparing data for display:{Style.RESET_ALL}")
    steps = 3  # Fewer steps
    for i in range(steps + 1):
        print_progress_bar(i, steps, 
                          prefix=f'{Fore.CYAN}Preparing:{Style.RESET_ALL}', 
                          suffix=f'Complete', 
                          print_end='\r')
        time.sleep(0.01)  # Much faster
        
    df = pd.DataFrame(candidates)
    df['rank'] = range(1, len(df) + 1)
    df = df[['rank', 'registration', 'name', 'p1_score', 'p1_correct', 'p2_score', 'p2_correct', 'final_score']]
    
    # Rename columns for better display
    df.columns = ['Rank', 'Registration', 'Name', 'P1 Score', 'P1 Correct', 'P2 Score', 'P2 Correct', 'Final Score']
    
    print(f"\n{Fore.GREEN}✓ Data preparation complete{Style.RESET_ALL}")
    
    # Format the header
    header = [
        f"\n{Fore.YELLOW}EMPRESA BRASILEIRA DE PESQUISA AGROPECUÁRIA (EMBRAPA){Style.RESET_ALL}",
        f"{Fore.YELLOW}CONCURSO PÚBLICO - EDITAL Nº 17 – EMBRAPA, DE 28 DE ABRIL DE 2025{Style.RESET_ALL}",
        f"\n{Fore.CYAN}OPÇÃO {option_number}: {option_title}{Style.RESET_ALL}\n",
        f"{Fore.GREEN}Total Candidates: {len(candidates)}{Style.RESET_ALL}\n"
    ]
    
    # Print header
    for line in header:
        print(line)
    
    # Generate table for display - Faster progress
    print(f"{Fore.YELLOW}⏳ Formatting table:{Style.RESET_ALL}")
    steps = 3
    for i in range(steps + 1):
        print_progress_bar(i, steps, 
                          prefix=f'{Fore.CYAN}Formatting:{Style.RESET_ALL}', 
                          suffix=f'Complete', 
                          print_end='\r')
        time.sleep(0.01)
    
    # Create the table
    table = tabulate(
        df, 
        headers='keys',
        tablefmt="grid",
        showindex=False
    )
    
    print(f"\n{Fore.GREEN}✓ Table formatting complete{Style.RESET_ALL}")
    
    # Print the table (show only first 20 rows)
    print(f"\n{Fore.YELLOW}Ranking Results (showing top 20):{Style.RESET_ALL}")
    print("\n".join(table.split("\n")[:22]))  # Table header + 20 rows
    print(f"\n{Fore.CYAN}... and {len(df)-20 if len(df)>20 else 0} more rows (full results in output file) ...{Style.RESET_ALL}\n")
    
    # Write to file if output_file is specified
    if output_file:
        try:
            # Determine file extension based on format
            if output_format.lower() == 'xlsx':
                if not output_file.lower().endswith('.xlsx'):
                    output_file = output_file.replace('.txt', '.xlsx').replace('.csv', '.xlsx')
                    if not output_file.lower().endswith('.xlsx'):
                        output_file += '.xlsx'
                
                print(f"{Fore.YELLOW}⏳ Saving ranking to Excel file...{Style.RESET_ALL}")
                
                # Show file writing progress - Faster progress
                steps = 5
                for i in range(steps + 1):
                    print_progress_bar(i, steps, 
                                      prefix=f'{Fore.CYAN}Saving Excel:{Style.RESET_ALL}', 
                                      suffix=f'Complete', 
                                      print_end='\r')
                    if i == 2:  # Write the file earlier
                        # Write to Excel with formatting
                        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                            format_excel(df, writer, option_number, option_title, len(candidates))
                    time.sleep(0.01)  # Faster progress
                
                print(f"\n{Fore.GREEN}✅ Ranking successfully saved to Excel file: {output_file}{Style.RESET_ALL}")
                
            elif output_format.lower() == 'csv':
                if not output_file.lower().endswith('.csv'):
                    output_file = output_file.replace('.txt', '.csv').replace('.xlsx', '.csv')
                    if not output_file.lower().endswith('.csv'):
                        output_file += '.csv'
                
                print(f"{Fore.YELLOW}⏳ Saving ranking to CSV file...{Style.RESET_ALL}")
                
                # Show file writing progress - Faster progress
                steps = 5
                for i in range(steps + 1):
                    print_progress_bar(i, steps, 
                                      prefix=f'{Fore.CYAN}Saving CSV:{Style.RESET_ALL}', 
                                      suffix=f'Complete', 
                                      print_end='\r')
                    if i == 2:  # Write the file earlier
                        # Write to CSV
                        df.to_csv(output_file, index=False, encoding='utf-8')
                    time.sleep(0.01)  # Faster progress
                
                print(f"\n{Fore.GREEN}✅ Ranking successfully saved to CSV file: {output_file}{Style.RESET_ALL}")
                
            else:  # Default to table format
                if not output_file.lower().endswith('.txt'):
                    output_file = output_file.replace('.csv', '.txt').replace('.xlsx', '.txt')
                    if not output_file.lower().endswith('.txt'):
                        output_file += '.txt'
                        
                print(f"{Fore.YELLOW}⏳ Saving ranking to text file...{Style.RESET_ALL}")
                
                # Show file writing progress - Faster progress
                steps = 5
                for i in range(steps + 1):
                    print_progress_bar(i, steps, 
                                      prefix=f'{Fore.CYAN}Saving Table:{Style.RESET_ALL}', 
                                      suffix=f'Complete', 
                                      print_end='\r')
                    if i == 2:  # Write the file earlier
                        # Remove color codes from header for file output
                        clean_header = [
                            "\nEMPRESA BRASILEIRA DE PESQUISA AGROPECUÁRIA (EMBRAPA)",
                            "CONCURSO PÚBLICO - EDITAL Nº 17 – EMBRAPA, DE 28 DE ABRIL DE 2025",
                            f"\nOPÇÃO {option_number}: {option_title}\n",
                            f"Total Candidates: {len(candidates)}\n"
                        ]
                        with open(output_file, 'w', encoding='utf-8') as f:
                            f.write("\n".join(clean_header))
                            f.write("\n")
                            f.write(table)
                    time.sleep(0.01)  # Faster progress
                
                print(f"\n{Fore.GREEN}✅ Ranking successfully saved to text file: {output_file}{Style.RESET_ALL}")
                
        except Exception as e:
            print(f"\n{Fore.RED}❌ Error writing to file: {e}{Style.RESET_ALL}")

def main():
    input_file = 'opcao_40000188.txt'
    output_file = 'ranking_40000188.xlsx'  # Changed to XLSX extension
    output_format = 'xlsx'  # Default to Excel format
    
    # Record start time for total execution
    total_start_time = time.time()
    
    # Print title with colors
    print(f"\n{Back.BLUE}{Fore.WHITE} EMBRAPA RANKING GENERATOR {Style.RESET_ALL}\n")
    print(f"{Fore.CYAN}Building ranking from file: {Fore.YELLOW}{input_file}{Style.RESET_ALL}")
    print(f"{Fore.CYAN}Output format: {Fore.YELLOW}{output_format.upper()}{Style.RESET_ALL}")
    print(f"{Fore.CYAN}{'=' * 60}{Style.RESET_ALL}")
    
    # Show overall progress
    steps = 3
    current_step = 0
    
    # Step 1: Parse candidate data
    print(f"\n{Back.CYAN}{Fore.BLACK} STEP 1: PARSING CANDIDATE DATA {Style.RESET_ALL}")
    start_time = time.time()
    option_number, option_title, candidates = parse_candidates_data(input_file)
    current_step += 1
    elapsed = time.time() - start_time
    print_progress_bar(current_step, steps, 
                      prefix=f'{Fore.BLUE}Overall Progress:{Style.RESET_ALL}', 
                      suffix=f'Step {current_step}/{steps} complete - Time: {elapsed:.2f}s', 
                      print_end='\n')
    print(f"{Fore.CYAN}{'=' * 60}{Style.RESET_ALL}")
    
    if candidates:
        # Step 2: Build ranking
        print(f"\n{Back.CYAN}{Fore.BLACK} STEP 2: BUILDING RANKING {Style.RESET_ALL}")
        start_time = time.time()
        ranked_candidates = build_ranking(candidates)
        current_step += 1
        elapsed = time.time() - start_time
        print_progress_bar(current_step, steps, 
                          prefix=f'{Fore.BLUE}Overall Progress:{Style.RESET_ALL}', 
                          suffix=f'Step {current_step}/{steps} complete - Time: {elapsed:.2f}s', 
                          print_end='\n')
        print(f"{Fore.CYAN}{'=' * 60}{Style.RESET_ALL}")
        
        # Step 3: Display and save ranking
        print(f"\n{Back.CYAN}{Fore.BLACK} STEP 3: GENERATING RESULTS {Style.RESET_ALL}")
        start_time = time.time()
        display_ranking(option_number, option_title, ranked_candidates, output_format, output_file)
        current_step += 1
        elapsed = time.time() - start_time
        print_progress_bar(current_step, steps, 
                          prefix=f'{Fore.BLUE}Overall Progress:{Style.RESET_ALL}', 
                          suffix=f'Step {current_step}/{steps} complete - Time: {elapsed:.2f}s', 
                          print_end='\n')
        print(f"{Fore.CYAN}{'=' * 60}{Style.RESET_ALL}")
        
        # Final message
        total_elapsed = time.time() - total_start_time
        print(f"\n{Back.GREEN}{Fore.BLACK} 🎉 RANKING PROCESS COMPLETED SUCCESSFULLY! {Style.RESET_ALL}")
        print(f"{Fore.GREEN}Total processing time: {total_elapsed:.2f} seconds{Style.RESET_ALL}")
        print(f"{Fore.GREEN}Output file: {output_file}{Style.RESET_ALL}")
        print(f"{Fore.GREEN}Candidates found: {len(candidates)}{Style.RESET_ALL}")
    else:
        print(f"{Back.RED}{Fore.WHITE} ❌ FAILED TO BUILD RANKING. PLEASE CHECK THE INPUT FILE. {Style.RESET_ALL}")

if __name__ == "__main__":
    main()